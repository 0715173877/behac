from django.core.mail import send_mail
from django.utils import timezone
from .models import Payment, FeeStructure, FeeReminder
from students.models import Student
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from .models import Payment, FeeCategory, FeeStructure
from .forms import PaymentForm, FeeCategoryForm, FeeStructureForm
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.shortcuts import get_object_or_404, render, redirect
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
from .forms import PaymentForm, FeeCategoryForm, PaymentBulkUploadForm
import uuid
from django.contrib import messages
from django.db import transaction
from django.views.generic import FormView
import openpyxl
from openpyxl.styles import Font
import pandas as pd
from academics.models import ClassLevel

class AdminRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_superuser or self.request.user.groups.filter(name__in=['Admin', 'Owner', 'Accountant']).exists()


from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from students.models import Student

@login_required
def select_student_for_payment(request):
    """Step 1: Search & select student, then redirect to payment form"""
    q = request.GET.get('q', '').strip()
    students = Student.objects.none()
    if q:
        students = Student.objects.filter(is_active=True).filter(
            Q(first_name__icontains=q) | Q(middle_name__icontains=q) |
            Q(last_name__icontains=q) | Q(registration_number__icontains=q) |
            Q(birth_cert_number__icontains=q)
        ).select_related('current_class').order_by('first_name')[:30]
    
    student_id = request.GET.get('student_id')
    if student_id:
        return redirect(reverse('finance:payment_create') + f'?student={student_id}')
    
    context = {
        'students': students,
        'query': q,
        'total_active': Student.objects.filter(is_active=True).count(),
    }
    return render(request, 'finance/select_student.html', context)

class AccountantDashboardView(LoginRequiredMixin, AdminRequiredMixin, ListView):
    model = Payment
    template_name = 'finance/accountant_dashboard.html'
    context_object_name = 'recent_payments'
    
    def get_queryset(self):
        return Payment.objects.all().select_related('student', 'fee_category').order_by('-payment_date')[:10]
    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        
        # Total income
        total_income = Payment.objects.aggregate(total=Sum('amount_paid'))['total'] or 0
        ctx['total_income'] = total_income
        
        # Income this month
        from django.utils import timezone
        now = timezone.now()
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        month_income = Payment.objects.filter(payment_date__gte=month_start).aggregate(total=Sum('amount_paid'))['total'] or 0
        ctx['month_income'] = month_income
        
        # Income today
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        today_income = Payment.objects.filter(payment_date__gte=today_start).aggregate(total=Sum('amount_paid'))['total'] or 0
        ctx['today_income'] = today_income
        
        # Total students with outstanding (FeeStructure total - Payments total per student)
        from academics.models import ClassLevel
        
        students = Student.objects.filter(is_active=True)
        ctx['total_students'] = students.count()
        
        # Students with no payments at all
        students_no_payment = students.filter(payments__isnull=True).count()
        ctx['students_no_payment'] = students_no_payment
        
        # Income by category
        income_by_category = Payment.objects.values('fee_category__name').annotate(total=Sum('amount_paid')).order_by('-total')
        ctx['income_by_category'] = income_by_category
        
        # Income by month (last 6 months)
        from datetime import datetime, timedelta
        six_months_ago = now - timedelta(days=180)
        income_by_month = []
        payments_qs = Payment.objects.filter(payment_date__gte=six_months_ago).order_by('payment_date')
        monthly_data = {}
        for p in payments_qs:
            key = p.payment_date.strftime('%Y-%m')
            if key not in monthly_data:
                monthly_data[key] = 0
            monthly_data[key] += p.amount_paid
        ctx['monthly_income'] = sorted(monthly_data.items())
        
        # Outstanding students list
        from django.db.models import OuterRef, Subquery
        
        # Total fee structures per class
        class_fees = FeeStructure.objects.values('class_level').annotate(total_fee=Sum('amount'))
        fee_by_class = {}
        for cf in class_fees:
            fee_by_class[cf['class_level']] = cf['total_fee']
        
        outstanding_students = []
        for s in students.select_related('current_class').prefetch_related('payments'):
            total_paid = sum(p.amount_paid for p in s.payments.all())
            expected_fee = fee_by_class.get(s.current_class_id, 0)
            balance = expected_fee - total_paid
            if balance > 0:
                outstanding_students.append({
                    'student': s,
                    'class_name': s.current_class.name if s.current_class else '-',
                    'expected': expected_fee,
                    'paid': total_paid,
                    'balance': balance
                })
        
        outstanding_students.sort(key=lambda x: x['balance'], reverse=True)
        ctx['outstanding_students'] = outstanding_students[:10]
        ctx['outstanding_count'] = len(outstanding_students)
        
        return ctx

from django.utils.dateparse import parse_date
from datetime import datetime, timedelta

@login_required
def export_payments_excel(request):
    """Export payments filtered by date range to Excel"""
    from_date = request.GET.get('from', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    to_date = request.GET.get('to', datetime.now().strftime('%Y-%m-%d'))
    category_id = request.GET.get('category', '')
    student_q = request.GET.get('student', '')
    
    payments = Payment.objects.filter(payment_date__gte=parse_date(from_date), payment_date__lte=parse_date(to_date))
    
    if category_id:
        payments = payments.filter(fee_category_id=category_id)
    if student_q:
        payments = payments.filter(
            Q(student__first_name__icontains=student_q) |
            Q(student__last_name__icontains=student_q) |
            Q(student__registration_number__icontains=student_q)
        )
    
    payments = payments.select_related('student', 'fee_category').order_by('-payment_date')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payments Report"
    
    headers = ['Date', 'Receipt #', 'Student', 'Reg Number', 'Class', 'Category', 'Amount', 'Month', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    for row, p in enumerate(payments, 2):
        ws.cell(row=row, column=1, value=p.payment_date.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=p.receipt_number)
        ws.cell(row=row, column=3, value=p.student.full_name)
        ws.cell(row=row, column=4, value=p.student.registration_number)
        ws.cell(row=row, column=5, value=p.student.current_class.name if p.student.current_class else '-')
        ws.cell(row=row, column=6, value=p.fee_category.name)
        ws.cell(row=row, column=7, value=float(p.amount_paid))
        ws.cell(row=row, column=8, value=p.installment_month or '-')
        ws.cell(row=row, column=9, value=p.notes or '')
    
    # Add totals row
    total_row = len(payments) + 2
    ws.cell(row=total_row, column=6, value='TOTAL:').font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=float(payments.aggregate(total=Sum('amount_paid'))['total'] or 0)).font = Font(bold=True)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="payments_report_{from_date}_to_{to_date}.xlsx"'
    wb.save(response)
    return response

@login_required
def export_outstanding_excel(request):
    """Export outstanding balances to Excel"""
    from django.db.models import Sum
    
    class_fees = FeeStructure.objects.values('class_level').annotate(total_fee=Sum('amount'))
    fee_by_class = {}
    for cf in class_fees:
        fee_by_class[cf['class_level']] = cf['total_fee']
    
    students = Student.objects.filter(is_active=True).select_related('current_class').prefetch_related('payments').order_by('first_name')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Outstanding Balances"
    
    headers = ['#', 'Student Name', 'Reg Number', 'Class', 'Expected Fee', 'Total Paid', 'Balance']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    row = 2
    for s in students:
        total_paid = sum(p.amount_paid for p in s.payments.all())
        expected_fee = fee_by_class.get(s.current_class_id, 0)
        balance = expected_fee - total_paid
        if balance > 0:
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=s.full_name)
            ws.cell(row=row, column=3, value=s.registration_number)
            ws.cell(row=row, column=4, value=s.current_class.name if s.current_class else '-')
            ws.cell(row=row, column=5, value=float(expected_fee))
            ws.cell(row=row, column=6, value=float(total_paid))
            ws.cell(row=row, column=7, value=float(balance))
            row += 1
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="outstanding_balances.xlsx"'
    wb.save(response)
    return response

@login_required
def outstanding_report(request):
    """Report of students with outstanding balances"""
    from django.db.models import Sum
    
    class_fees = FeeStructure.objects.values('class_level').annotate(total_fee=Sum('amount'))
    fee_by_class = {}
    for cf in class_fees:
        fee_by_class[cf['class_level']] = cf['total_fee']
    
    students = Student.objects.filter(is_active=True).select_related('current_class').prefetch_related('payments').order_by('first_name')
    
    outstanding_list = []
    cleared_count = 0
    total_expected = 0
    total_paid_all = 0
    total_balance = 0
    
    for s in students:
        total_paid = sum(p.amount_paid for p in s.payments.all())
        expected_fee = fee_by_class.get(s.current_class_id, 0)
        balance = expected_fee - total_paid
        total_expected += expected_fee
        total_paid_all += total_paid
        if balance > 0:
            total_balance += balance
            outstanding_list.append({
                'student': s,
                'class_name': s.current_class.name if s.current_class else '-',
                'class_id': s.current_class_id,
                'expected': expected_fee,
                'paid': total_paid,
                'balance': balance
            })
        else:
            cleared_count += 1
    
    outstanding_list.sort(key=lambda x: x['balance'], reverse=True)
    
    # Filter by class if specified
    class_filter = request.GET.get('class')
    if class_filter:
        outstanding_list = [o for o in outstanding_list if o['class_id'] == int(class_filter)]
    
    # Filter by min balance
    min_balance = request.GET.get('min_balance', 0)
    try:
        min_balance = float(min_balance)
        if min_balance > 0:
            outstanding_list = [o for o in outstanding_list if o['balance'] >= min_balance]
    except ValueError:
        pass
    
    classes = ClassLevel.objects.all()
    
    context = {
        'outstanding_list': outstanding_list,
        'cleared_count': cleared_count,
        'total_students': students.count(),
        'total_expected': total_expected,
        'total_paid': total_paid_all,
        'total_balance': total_balance,
        'classes': classes,
        'class_filter': class_filter,
    }
    return render(request, 'finance/outstanding_report.html', context)

@login_required
def payment_report(request):
    """Payment report page with filters"""
    from_date = request.GET.get('from', (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    to_date = request.GET.get('to', datetime.now().strftime('%Y-%m-%d'))
    category_id = request.GET.get('category', '')
    
    payments = Payment.objects.filter(payment_date__gte=parse_date(from_date), payment_date__lte=parse_date(to_date))
    
    if category_id:
        payments = payments.filter(fee_category_id=category_id)
    
    payments = payments.select_related('student', 'fee_category').order_by('-payment_date')
    
    total = payments.aggregate(total=Sum('amount_paid'))['total'] or 0
    categories = FeeCategory.objects.all()
    
    context = {
        'payments': payments,
        'from_date': from_date,
        'to_date': to_date,
        'selected_category': int(category_id) if category_id else None,
        'total': total,
        'categories': categories,
    }
    return render(request, 'finance/payment_report.html', context)

def download_payment_template(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Payments"
    columns = ['birth_cert_number', 'fee_category', 'amount', 'payment_date', 'installment_month', 'notes']
    sheet.append(columns)
    for col in sheet.iter_cols(max_row=1, max_col=len(columns)):
        col[0].font = Font(bold=True)
    sheet.append(['BC12345', 'School Fee', 50000, '2025-03-15', 3, 'March tuition'])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="payment_import_template.xlsx"'
    workbook.save(response)
    return response

# =========== PAYMENT VIEWS ===========

class PaymentListView(LoginRequiredMixin, ListView):
    model = Payment
    template_name = 'finance/payment_list.html'
    context_object_name = 'payments'
    paginate_by = 50

    def get_queryset(self):
        qs = super().get_queryset()
        q = self.request.GET.get('q', '')
        cat = self.request.GET.get('category', '')
        if q:
            qs = qs.filter(student__first_name__icontains=q) | qs.filter(student__last_name__icontains=q) | qs.filter(receipt_number__icontains=q)
        if cat:
            qs = qs.filter(fee_category_id=cat)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['categories'] = FeeCategory.objects.all()
        ctx['q'] = self.request.GET.get('q', '')
        ctx['selected_category'] = self.request.GET.get('category', '')
        return ctx

class PaymentCreateView(LoginRequiredMixin, AdminRequiredMixin, CreateView):
    model = Payment
    form_class = PaymentForm
    template_name = 'finance/payment_form.html'
    success_url = reverse_lazy('finance:payment_list')

    def get_initial(self):
        initial = super().get_initial()
        student_id = self.request.GET.get('student')
        if student_id:
            try:
                student = Student.objects.get(pk=student_id)
                initial['student'] = student
            except Student.DoesNotExist:
                pass
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['all_students'] = Student.objects.filter(is_active=True).select_related('current_class').order_by('first_name')
        # Pass selected student to template
        student_id = self.request.GET.get('student')
        if student_id:
            try:
                ctx['selected_student'] = Student.objects.get(pk=student_id)
            except Student.DoesNotExist:
                pass
        return ctx

    def form_valid(self, form):
        form.instance.receipt_number = str(uuid.uuid4()).replace('-', '').upper()[:12]
        return super().form_valid(form)

class PaymentUpdateView(LoginRequiredMixin, AdminRequiredMixin, UpdateView):
    model = Payment
    form_class = PaymentForm
    template_name = 'finance/payment_form.html'
    success_url = reverse_lazy('finance:payment_list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['all_students'] = Student.objects.filter(is_active=True).select_related('current_class').order_by('first_name')
        return ctx

class PaymentDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = Payment
    template_name = 'finance/payment_confirm_delete.html'
    success_url = reverse_lazy('finance:payment_list')

class ReceiptView(LoginRequiredMixin, AdminRequiredMixin, DetailView):
    model = Payment
    template_name = 'finance/receipt.html'

def payment_receipt_pdf(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    html_string = render_to_string('finance/receipt_pdf.html', {'payment': payment})
    pdf = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="receipt_{payment.receipt_number}.pdf"'
    return response

class PaymentBulkUploadView(LoginRequiredMixin, AdminRequiredMixin, FormView):
    template_name = 'finance/payment_bulk_upload.html'
    form_class = PaymentBulkUploadForm
    success_url = reverse_lazy('finance:payment_list')

    def form_valid(self, form):
        excel_file = self.request.FILES['excel_file']
        try:
            df = pd.read_excel(excel_file)
            required_columns = ['birth_cert_number', 'fee_category', 'amount', 'payment_date']
            missing = [col for col in required_columns if col not in df.columns]
            if missing:
                raise Exception(f"Missing columns: {missing}")

            with transaction.atomic():
                success_count = 0
                error_rows = []

                for idx, row in df.iterrows():
                    row_dict = {k: (str(v).strip() if pd.notna(v) else '') for k, v in row.items()}
                    birth_cert = row_dict.get('birth_cert_number')
                    if not birth_cert:
                        error_rows.append((idx+2, "Missing birth certificate number"))
                        continue

                    try:
                        student = Student.objects.get(birth_cert_number=birth_cert)
                    except Student.DoesNotExist:
                        error_rows.append((idx+2, f"Student with birth cert '{birth_cert}' not found"))
                        continue

                    fee_cat_name = row_dict.get('fee_category')
                    if not fee_cat_name:
                        error_rows.append((idx+2, "Missing fee category"))
                        continue
                    try:
                        fee_category = FeeCategory.objects.get(name=fee_cat_name)
                    except FeeCategory.DoesNotExist:
                        error_rows.append((idx+2, f"Fee category '{fee_cat_name}' not found"))
                        continue

                    try:
                        amount = float(row_dict.get('amount', 0))
                    except:
                        error_rows.append((idx+2, "Invalid amount"))
                        continue

                    payment_date_str = row_dict.get('payment_date')
                    if not payment_date_str:
                        error_rows.append((idx+2, "Missing payment date"))
                        continue
                    try:
                        from datetime import datetime
                        payment_date = datetime.strptime(payment_date_str, '%Y-%m-%d').date()
                    except:
                        error_rows.append((idx+2, "Invalid payment date (use YYYY-MM-DD)"))
                        continue

                    installment = row_dict.get('installment_month', '')
                    installment_month = int(installment) if installment and installment.isdigit() else None
                    notes = row_dict.get('notes', '')

                    receipt_number = str(uuid.uuid4()).replace('-', '').upper()[:12]
                    Payment.objects.create(
                        student=student,
                        fee_category=fee_category,
                        amount_paid=amount,
                        payment_date=payment_date,
                        receipt_number=receipt_number,
                        installment_month=installment_month,
                        notes=notes
                    )
                    success_count += 1

                if error_rows:
                    self.request.session['payment_bulk_errors'] = error_rows
                    messages.warning(self.request, f"Imported {success_count} payments with {len(error_rows)} errors.")
                else:
                    messages.success(self.request, f"Successfully imported {success_count} payments.")

        except Exception as e:
            messages.error(self.request, f"Import failed: {str(e)}")
            return self.form_invalid(form)

        return super().form_valid(form)

# =========== FEE CATEGORY VIEWS ===========

class FeeCategoryListView(LoginRequiredMixin, ListView):
    model = FeeCategory
    template_name = 'finance/category_list.html'
    context_object_name = 'categories'

class FeeCategoryCreateView(LoginRequiredMixin, AdminRequiredMixin, CreateView):
    model = FeeCategory
    form_class = FeeCategoryForm
    template_name = 'finance/category_form.html'
    success_url = reverse_lazy('finance:category_list')

class FeeCategoryUpdateView(LoginRequiredMixin, AdminRequiredMixin, UpdateView):
    model = FeeCategory
    form_class = FeeCategoryForm
    template_name = 'finance/category_form.html'
    success_url = reverse_lazy('finance:category_list')

class FeeCategoryDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = FeeCategory
    template_name = 'finance/category_confirm_delete.html'
    success_url = reverse_lazy('finance:category_list')

# =========== FEE STRUCTURE VIEWS ===========

class FeeStructureListView(LoginRequiredMixin, ListView):
    model = FeeStructure
    template_name = 'finance/feestructure_list.html'
    context_object_name = 'feestructures'

    def get_queryset(self):
        qs = super().get_queryset()
        q = self.request.GET.get('q', '')
        cls = self.request.GET.get('class', '')
        cat = self.request.GET.get('category', '')
        if q:
            qs = qs.filter(fee_category__name__icontains=q)
        if cls:
            qs = qs.filter(class_level_id=cls)
        if cat:
            qs = qs.filter(fee_category_id=cat)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['classlevels'] = ClassLevel.objects.all()
        ctx['categories'] = FeeCategory.objects.all()
        ctx['q'] = self.request.GET.get('q', '')
        ctx['selected_class'] = self.request.GET.get('class', '')
        ctx['selected_category'] = self.request.GET.get('category', '')
        return ctx

class FeeStructureCreateView(LoginRequiredMixin, AdminRequiredMixin, CreateView):
    model = FeeStructure
    form_class = FeeStructureForm
    template_name = 'finance/feestructure_form.html'
    success_url = reverse_lazy('finance:feestructure_list')

class FeeStructureUpdateView(LoginRequiredMixin, AdminRequiredMixin, UpdateView):
    model = FeeStructure
    form_class = FeeStructureForm
    template_name = 'finance/feestructure_form.html'
    success_url = reverse_lazy('finance:feestructure_list')

class FeeStructureDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = FeeStructure
    template_name = 'finance/feestructure_confirm_delete.html'
    success_url = reverse_lazy('finance:feestructure_list')

# =========== FEE REMINDERS ===========

def send_fee_reminders():
    today = timezone.now().date()
    students = Student.objects.filter(is_active=True)
    for s in students:
        total_due = FeeStructure.objects.filter(class_level=s.current_class).aggregate(total=Sum('amount'))['total'] or 0
        total_paid = s.payments.aggregate(total=Sum('amount_paid'))['total'] or 0
        due = total_due - total_paid
        if due > 0:
            send_mail(
                subject='Fee Reminder - Behac International Academy',
                message=f'Dear Parent, your child {s.full_name} has an outstanding fee of {due}. Please clear by end of month.',
                from_email='finance@behac.academy',
                recipient_list=[s.parent_mobile + '@smsgateway.com', s.parent_user.email],
                fail_silently=True,
            )
            FeeReminder.objects.create(student=s, due_date=today, amount_due=due, sent_via='email')
